"""Safe writes to the existing SkillSignalZA workbook template."""

from __future__ import annotations

import os
import re
import shutil
import threading
from copy import copy
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook

from .schema import TRACK_CODES, TRACK_HEADERS


class WorkbookError(RuntimeError):
    """Raised when a record cannot be validated or written."""


class WorkbookStore:
    def __init__(self, template_path: Path, working_path: Path) -> None:
        self.template_path = Path(template_path)
        self.working_path = Path(working_path)
        self.backup_path = self.working_path.with_name(self.working_path.stem + "_LastBackup.xlsx")
        self._lock = threading.Lock()

    def ensure_working_copy(self) -> Path:
        if not self.template_path.exists():
            raise WorkbookError(f"Workbook template not found: {self.template_path.name}")
        self.working_path.parent.mkdir(parents=True, exist_ok=True)
        if not self.working_path.exists():
            shutil.copy2(self.template_path, self.working_path)
            workbook = load_workbook(self.working_path)
            try:
                for track in TRACK_HEADERS:
                    sheet = workbook[track]
                    if str(sheet.cell(2, 1).value or "").startswith("EXAMPLE"):
                        for column in range(1, 22):
                            sheet.cell(2, column).value = None
                workbook.save(self.working_path)
            finally:
                workbook.close()
        return self.working_path

    def status(self) -> dict[str, object]:
        path = self.ensure_working_copy()
        with self._lock:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                counts: dict[str, int] = {}
                recent: list[dict[str, str]] = []
                for track in TRACK_HEADERS:
                    sheet = workbook[track]
                    headers = [str(cell.value or "") for cell in sheet[1]]
                    positions = {name: index for index, name in enumerate(headers)}
                    records: list[dict[str, str]] = []
                    for values in sheet.iter_rows(min_row=2, values_only=True):
                        post_id = values[positions["Post ID"]]
                        if not post_id or str(post_id).startswith("EXAMPLE"):
                            continue
                        records.append(
                            {
                                "id": str(post_id),
                                "track": TRACK_CODES[track],
                                "title": str(values[positions["Job Title (as written)"]] or ""),
                                "company": str(values[positions["Company Name"]] or ""),
                            }
                        )
                    counts[track] = len(records)
                    recent.extend(records[-3:])
                recent.sort(key=lambda item: _id_number(item["id"]), reverse=True)
                return {
                    "counts": counts,
                    "recent": recent[:6],
                    "export_name": self.working_path.name,
                }
            finally:
                workbook.close()

    def add_record(self, track: str, record: dict[str, object]) -> dict[str, str]:
        if track not in TRACK_HEADERS:
            raise WorkbookError("Select either SE Track or DA Track.")
        cleaned = self._validate_record(track, record)
        path = self.ensure_working_copy()
        temp_path = path.with_name(path.stem + ".tmp.xlsx")

        with self._lock:
            workbook = load_workbook(path)
            try:
                sheet = workbook[track]
                headers = [str(cell.value or "") for cell in sheet[1]]
                if headers[: len(TRACK_HEADERS[track])] != TRACK_HEADERS[track]:
                    raise WorkbookError(f"The {track} headers no longer match the expected 21-column template.")
                self._reject_duplicate(sheet, headers, cleaned)
                post_id = self._next_id(sheet, TRACK_CODES[track])
                row_number = self._next_row(sheet)
                if row_number > sheet.max_row:
                    sheet.insert_rows(row_number)
                    self._copy_row_style(sheet, row_number - 1, row_number)
                cleaned["Post ID"] = post_id
                for column, header in enumerate(headers[:21], start=1):
                    sheet.cell(row=row_number, column=column).value = _safe_excel_value(cleaned.get(header, ""))

                if path.exists():
                    shutil.copy2(path, self.backup_path)
                workbook.save(temp_path)
                os.replace(temp_path, path)
                return {"post_id": post_id, "track": track, "path": str(path)}
            except PermissionError as exc:
                raise WorkbookError(
                    "Excel appears to have the collection workbook open. Close it, then try Create record again."
                ) from exc
            finally:
                workbook.close()
                if temp_path.exists():
                    temp_path.unlink(missing_ok=True)

    def _validate_record(self, track: str, record: dict[str, object]) -> dict[str, str]:
        cleaned = {header: str(record.get(header, "") or "").strip() for header in TRACK_HEADERS[track]}
        url = cleaned["Job Post URL"]
        if not url.startswith(("http://", "https://")):
            raise WorkbookError("Job Post URL must start with http:// or https://.")
        flag_header = (
            "Git/GitHub Explicitly Named (Y/N)"
            if track == "SE Track"
            else "SQL Explicitly Named (Y/N)"
        )
        for header in [flag_header, "Portfolio/GitHub Requested (Y/N)"]:
            cleaned[header] = cleaned[header].upper()
            if cleaned[header] not in {"Y", "N"}:
                raise WorkbookError(f"{header} must contain Y or N.")
        level_header = "Requirement Level (Required/Preferred/Not Mentioned)"
        valid_levels = {"Required", "Preferred", "Not Mentioned"}
        if cleaned[level_header] not in valid_levels:
            raise WorkbookError(f"{level_header} must be Required, Preferred, or Not Mentioned.")
        if cleaned["Portfolio/GitHub Requested (Y/N)"] == "N":
            cleaned[level_header] = "Not Mentioned"
        return cleaned

    @staticmethod
    def _reject_duplicate(sheet, headers: list[str], record: dict[str, str]) -> None:
        url_col = headers.index("Job Post URL") + 1
        title_col = headers.index("Job Title (as written)") + 1
        company_col = headers.index("Company Name") + 1
        location_col = headers.index("Location") + 1
        target_url = _normalized_url(record["Job Post URL"])
        target_signature = tuple(
            record[name].casefold().strip()
            for name in ["Job Title (as written)", "Company Name", "Location"]
        )
        for row in range(2, sheet.max_row + 1):
            existing_url = sheet.cell(row, url_col).value
            if existing_url and _normalized_url(str(existing_url)) == target_url:
                raise WorkbookError("This exact job-post URL is already in the selected track.")
            signature = tuple(
                str(sheet.cell(row, col).value or "").casefold().strip()
                for col in [title_col, company_col, location_col]
            )
            if all(target_signature) and signature == target_signature:
                raise WorkbookError("A record with the same title, company, and location already exists.")

    @staticmethod
    def _next_id(sheet, prefix: str) -> str:
        values: list[int] = []
        pattern = re.compile(rf"^{re.escape(prefix)}-(\d+)$", re.I)
        for row in range(2, sheet.max_row + 1):
            value = str(sheet.cell(row, 1).value or "")
            match = pattern.match(value)
            if match:
                values.append(int(match.group(1)))
        return f"{prefix}-{(max(values, default=0) + 1):03d}"

    @staticmethod
    def _next_row(sheet) -> int:
        example_row = str(sheet.cell(2, 1).value or "")
        if example_row.startswith("EXAMPLE"):
            return 2
        for row in range(2, sheet.max_row + 1):
            if not sheet.cell(row, 1).value:
                return row
        return sheet.max_row + 1

    @staticmethod
    def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
        for column in range(1, sheet.max_column + 1):
            source = sheet.cell(source_row, column)
            target = sheet.cell(target_row, column)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            target.alignment = copy(source.alignment)
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def _normalized_url(value: str) -> str:
    parts = urlsplit(value.strip())
    path = parts.path.rstrip("/") or "/"
    return urlunsplit((parts.scheme.lower(), parts.netloc.lower(), path, parts.query, ""))


def _safe_excel_value(value: str) -> str:
    if value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _id_number(value: str) -> int:
    match = re.search(r"(\d+)$", value)
    return int(match.group(1)) if match else 0
